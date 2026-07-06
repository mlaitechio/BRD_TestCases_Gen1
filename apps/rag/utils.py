"""
RAG Utility Functions - Document extraction, chunking, embedding
"""

import logging
import os
import json
from typing import List, Dict

logger = logging.getLogger(__name__)


def extract_text_from_document(file_path: str, file_type: str) -> str:
    """
    Extract text from uploaded document (PDF, DOCX, TXT, XLSX, XLS, CSV).

    Args:
        file_path: Path to the document file
        file_type: File type ('pdf', 'docx', 'txt', 'xlsx', 'xls', 'csv')

    Returns:
        Extracted text content
    """
    if file_type == 'pdf':
        return _extract_from_pdf(file_path)
    elif file_type == 'docx':
        return _extract_from_docx(file_path)
    elif file_type == 'txt':
        return _extract_from_txt(file_path)
    elif file_type in ('xlsx', 'xls'):
        return _extract_from_excel(file_path)
    elif file_type == 'csv':
        return _extract_from_csv(file_path)
    else:
        raise ValueError(f'Unsupported file type: {file_type}')


def _extract_from_pdf(file_path: str) -> str:
    """Extract text from PDF"""
    try:
        import pdfplumber
        text = []
        with pdfplumber.open(file_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                page_text = page.extract_text()
                if page_text:
                    text.append(f'--- Page {page_num} ---\n{page_text}')
        return '\n\n'.join(text)
    except Exception as e:
        logger.error(f'PDF extraction failed: {e}')
        return ''


def _extract_from_docx(file_path: str) -> str:
    """Extract text from DOCX"""
    try:
        from docx import Document
        doc = Document(file_path)
        text = []
        for para in doc.paragraphs:
            if para.text.strip():
                text.append(para.text)
        return '\n'.join(text)
    except Exception as e:
        logger.error(f'DOCX extraction failed: {e}')
        return ''


def _extract_from_txt(file_path: str) -> str:
    """Extract text from TXT"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        logger.error(f'TXT extraction failed: {e}')
        return ''


def _extract_from_excel(file_path: str) -> str:
    """Extract text from Excel (XLSX, XLS)"""
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file_path)
        text_parts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"--- Sheet: {sheet_name} ---")

            for row in sheet.iter_rows(values_only=True):
                row_text = ' | '.join(str(cell) if cell is not None else '' for cell in row)
                if row_text.strip():
                    text_parts.append(row_text)

        return '\n'.join(text_parts)
    except Exception as e:
        logger.error(f'Excel extraction failed: {e}')
        return ''


def _extract_from_csv(file_path: str) -> str:
    """Extract text from CSV"""
    try:
        import csv
        text_parts = []

        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row_num, row in enumerate(reader, 1):
                row_text = ' | '.join(str(cell).strip() for cell in row if cell)
                if row_text.strip():
                    text_parts.append(row_text)

        return '\n'.join(text_parts)
    except Exception as e:
        logger.error(f'CSV extraction failed: {e}')
        return ''


def chunk_document(
    text: str,
    document_id: str,
    category: str = 'General',
    max_chars: int = 2000
) -> List[Dict]:
    """
    Split document text into chunks for embedding.

    Args:
        text: Document text
        document_id: Document UUID
        category: Document category
        max_chars: Max characters per chunk

    Returns:
        List of chunk dicts: [{'id': '...', 'text': '...', 'metadata': {...}}, ...]
    """
    if not text or len(text.strip()) < 50:
        return []

    chunks = []
    chunk_counter = 0

    # Split by paragraphs first
    paragraphs = text.split('\n\n')
    current_chunk = ''

    for para in paragraphs:
        if len(current_chunk) + len(para) + 2 <= max_chars:
            current_chunk += para + '\n\n'
        else:
            if current_chunk.strip():
                chunk_id = f'doc_{document_id}_chunk_{chunk_counter}'
                chunks.append({
                    'id': chunk_id,
                    'text': current_chunk.strip(),
                    'metadata': {
                        'document_id': str(document_id),
                        'category': category,
                        'chunk_index': chunk_counter
                    }
                })
                chunk_counter += 1
            current_chunk = para + '\n\n'

    # Add remaining chunk
    if current_chunk.strip():
        chunk_id = f'doc_{document_id}_chunk_{chunk_counter}'
        chunks.append({
            'id': chunk_id,
            'text': current_chunk.strip(),
            'metadata': {
                'document_id': str(document_id),
                'category': category,
                'chunk_index': chunk_counter
            }
        })

    logger.info(f'Created {len(chunks)} chunks for document {document_id}')
    return chunks


def format_rag_context(search_results: List[Dict]) -> str:
    """
    Format RAG search results into a context block for LLM injection.

    Args:
        search_results: List of search result dicts

    Returns:
        Formatted context string
    """
    if not search_results:
        return ''

    context_blocks = []
    for result in search_results:
        title = result.get('title', 'Unknown')
        content = result.get('content', '')
        score = result.get('similarity_score', 0)
        source = result.get('source', 'Unknown')

        block = f"""--- Retrieved from Global RAG [{source}] (Relevance: {score:.2f}) ---
Title: {title}

{content}
"""
        context_blocks.append(block)

    return '\n\n'.join(context_blocks)
