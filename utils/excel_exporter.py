import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def export_testcases_to_excel(test_case_data: dict) -> bytes:
    """
    Export structured Test Case JSON to an Excel (.xlsx) file.
    """
    wb = openpyxl.Workbook()
    
    # 1. Test Cases Sheet
    ws1 = wb.active
    ws1.title = "Test Cases"
    
    headers = [
        "SR No", "File Name", "Product Name", "Process Category", "BRD/FSD", 
        "Business Process ID", "Business Process", "BRD/FSD Reference", 
        "Scenario ID", "Scenario Description", "Category", "Importance", 
        "Test Case ID", "Creation Date", "Prepared By", "TC Module", 
        "TC Sub-Module", "Path", "Test Condition", "Pre-requisite", 
        "Test Case Description", "Test Priority", "Test Classification", 
        "Test Category", "Test Data", "Expected Result", "Actual Result", 
        "Release", "Execution Status", "Execution Date", "Executed By", 
        "Execution Result", "Defect Id", "Severity", "Priority", 
        "Defect Status", "Remarks", "Frequency(For Uat Purpose)", 
        "ABFL-IT Remarks", "OwnerShip"
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    row_num = 2
    for tc in test_case_data.get("test_cases", []):
        ws1.append([
            tc.get("sr_no", ""),
            tc.get("file_name", ""),
            tc.get("product_name", ""),
            tc.get("process_category", ""),
            tc.get("brd_fsd", ""),
            tc.get("business_process_id", ""),
            tc.get("business_process", ""),
            tc.get("brd_fsd_reference", ""),
            tc.get("scenario_id", ""),
            tc.get("scenario_description", ""),
            tc.get("category", ""),
            tc.get("importance", ""),
            tc.get("test_case_id", ""),
            tc.get("creation_date", ""),
            tc.get("prepared_by", ""),
            tc.get("tc_module", ""),
            tc.get("tc_sub_module", ""),
            tc.get("path", ""),
            tc.get("test_condition", ""),
            tc.get("pre_requisite", ""),
            tc.get("test_case_description", ""),
            tc.get("test_priority", ""),
            tc.get("test_classification", ""),
            tc.get("test_category", ""),
            tc.get("test_data", ""),
            tc.get("expected_result", ""),
            tc.get("actual_result", ""),
            tc.get("release", ""),
            tc.get("execution_status", ""),
            tc.get("execution_date", ""),
            tc.get("executed_by", ""),
            tc.get("execution_result", ""),
            tc.get("defect_id", ""),
            tc.get("severity", ""),
            tc.get("priority", ""),
            tc.get("defect_status", ""),
            tc.get("remarks", ""),
            tc.get("frequency", ""),
            tc.get("abfl_it_remarks", ""),
            tc.get("ownership", "")
        ])
        row_num += 1

    # 2. Traceability Matrix Sheet
    ws2 = wb.create_sheet(title="Traceability Matrix")
    headers_tm = ["Requirement ID", "Requirement Title", "Linked Test Cases", "Coverage Status"]
    for col_num, header in enumerate(headers_tm, 1):
        cell = ws2.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        
    for tm in test_case_data.get("traceability_matrix", []):
        ws2.append([
            tm.get("requirement_id", ""),
            tm.get("requirement_title", ""),
            ", ".join(tm.get("linked_test_cases", [])),
            tm.get("coverage_status", "")
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
